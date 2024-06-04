# -*- coding: utf-8 -*-
from email.mime import base
import requests
import bs4
import time
from sys import argv
from openpyxl import load_workbook
import json

def get_request(url):
    """
    HTTP request to store html as an object
    """
    res = requests.get(url, verify=False)
    if res.status_code == 200:
        global parse_soup
        parse_soup = bs4.BeautifulSoup(res.text, "html.parser")
    elif res.status_code == 404:
        print("The Status Code returned is", res.status_code, url)
        mega_menu_404.append(url)
    else:
        print("The Status Code returned is", res.status_code, url)
####
####
def navigation_unfiltered():
    """
        Stored list of navigation links from the mega menu // unfiltered
    """
    local_list = parse_soup.select('.list-unstyled.category-list a')
    for i in local_list:
        navigation_list_unfiltered.append(i.get("href"))
        nav_info = {'url': i.get("href"), 'navid': i.get("data-nav")}
        custom_data_dict.append(nav_info)
####
####
def navigation_duplicates(list_arg):
    """
        Function to remove duplicates from a list, preserves original list order
    """
    """ Removes URLs start with &, these are JS injected """
    [list_purge_duplicates_one.append(i['url'])
    for i in list_arg if not i['url'].startswith('&')]
    # print(list_purge_duplicates_one)
    """ Removes tracking """
    for i in list_purge_duplicates_one:
        # print(i)
        if '?' in i:
            x = i.split('?')[0]
            list_purge_duplicates_two.append(x)
            # print(list_purge_duplicates_two)
        elif '?' not in i:
            list_purge_duplicates_two.append(i)
    """ Removes duplicate URLS """
    [list_purge_duplicates.append(i) for i in list_purge_duplicates_two if not list_purge_duplicates.count(i)]
    
    return list_purge_duplicates
####
####
def remove_urls():
    """
        Filter out external urls such as blog/pinterest
    """
    global purge_external_urls
    
    purge_external_urls = list(list_purge_duplicates)
    [purge_external_urls.remove(i)
     for i in list_purge_duplicates if i.__contains__('.')]
    """ Removes empty URLs eg those from updateSaleLinks"""
    [purge_external_urls.remove(i)
     for i in list_purge_duplicates if i == ('')]
####
####
def domain_prefix_url():
    """
        Select the domain url based on user input from options banner
    """
    url_dict = {
    'UK-staging': 'https://uk-staging-p3-joules.eclipsegroup.co.uk',
    'UK-live': 'https://www.joules.com',
    'US-staging': 'https://us-staging-p3-joules.eclipsegroup.co.uk',
    'US-live': 'https://www.joulesusa.com',
    'DE-staging': 'https://de-staging-p3-joules.eclipsegroup.co.uk',
    'DE-live': 'https://www.tomjoule.de'
    }
    name_list = []
    for name_key in url_dict.keys():
        name_list.append(name_key)
    url_list = []
    for url_key in url_dict.values():
        url_list.append(url_key)
    return name_list, url_list
####
####
def nav_data_get(dict_arg):
    global navid_data
    navid_data = []
    for i in dict_arg:
        navid_data.append(i['navid'])
####
####
def category(domain_prefix_url, domain_suffix_url, domain_name_string, impex_string, navid_data):
    # workbook
    wb = load_workbook('Template.xlsx')
    # worksheet, get active
    ws = wb.active
    counter = 0
    create_dict = {'url': domain_suffix_url, 'navid': navid_data}
    for (category_url, navid) in zip(create_dict['url'], create_dict['navid']):
        html_doc = requests.get(str(domain_prefix_url) + str(category_url).strip(),
                                verify=False,
                                allow_redirects=False)
        soup = bs4.BeautifulSoup(html_doc.text, "html.parser")
        try:
            if category_url:
                pagination_data = soup.find("script", {"data-type": "pagination"})
                pagination_dataJson = json.loads(pagination_data.string)
                product_count = pagination_dataJson["totalNumberOfResults"]
                category_info = []
                if product_count < 5:
                   
                    print(f'{domain_prefix_url}{category_url} ~ Total Number of products in category: {product_count}')     
                    if 'None' in navid:
                        print('Null Nav ID')
                    else:
                        counter = counter + 1 
                        concat_nav_id = f';{navid}; " FALSE ";'
                        category_info.extend((category_url, product_count))
                    with open(f'./impex-header-templates/{impex_string}', 'a') as file:
                        file.write(concat_nav_id)
                        file.write('\n')
                    
                else:
                    print(f'{domain_prefix_url}{category_url} ~ Category not empty')
                for col, val in enumerate(category_info, start=1):
                    ws.cell(row=counter + 2, column=col).value = val
            else:
                break
        except:
            pass
    wb.save(filename=f'{domain_name_string}.xlsx')
    
####
####

"""selects the domain url from users choice"""
domain_func = domain_prefix_url()
domain_url = domain_func[1]
for domain_suffix_url in domain_url:
    custom_data_dict = []
    mega_menu_404, navigation_list_unfiltered, list_purge_duplicates = [], [], []
    list_purge_duplicates_one = []
    list_purge_duplicates_two = []
    print(f'\nUsing {domain_suffix_url} as the domain url\n')
    get_request(domain_suffix_url)
    navigation_unfiltered()
    navigation_duplicates(list_arg=custom_data_dict)
    nav_data_get(dict_arg=custom_data_dict)
    remove_urls()

    if 'uk-staging-p3' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'UK-staging', impex_string=f'UK-staging.impex', navid_data=navid_data)
    if 'www.joules.com' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'UK-live', impex_string=f'UK-live.impex', navid_data=navid_data)
    if 'us-staging-p3' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'US-staging', impex_string=f'US-staging.impex', navid_data=navid_data)
    if 'www.joulesusa.com' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'US-live', impex_string=f'US-live.impex', navid_data=navid_data)
    if 'de-staging-p3' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'DE-staging', impex_string=f'DE-staging.impex', navid_data=navid_data)
    if 'www.tomjoule.de' in domain_suffix_url:
        category(domain_prefix_url=domain_suffix_url, domain_suffix_url=purge_external_urls, domain_name_string=f'DE-live', impex_string=f'DE-live.impex', navid_data=navid_data)
    end = time.time()
    print(f'\nThe script took {end - start} seconds to run')
####
####
